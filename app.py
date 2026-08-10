import streamlit as st
import pandas as pd
import io
import json
import os
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
import plotly.express as px
from jinja2 import Template
from datetime import datetime
import re

CONFIG_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'config.json')

DEFAULT_CONFIG = {
    "dispatch_keywords": [],
    "import2": [],
    "export_tasks": [],
    "export_companies": [],
    "cost_categories": [],
    "import3_receivers": [],
    "consulting_tasks": [],
    "export_to_import3_registrars": [],
    "export_to_import2_registrars": []
}


def load_json_config():
    if os.path.exists(CONFIG_PATH):
        with open(CONFIG_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    try:
        if 'CONFIG_JSON' in st.secrets:
            return json.loads(st.secrets['CONFIG_JSON'])
    except Exception:
        pass
    return DEFAULT_CONFIG.copy()


def is_config_empty(config):
    return all(len(v) == 0 for v in config.values() if isinstance(v, list))


def save_json_config(config):
    with open(CONFIG_PATH, 'w', encoding='utf-8') as f:
        json.dump(config, f, ensure_ascii=False, indent=2)


# 사이드바 편집 위젯 key 목록 — 설정을 외부에서 갈아끼울 때 위젯 상태를 리셋해야
# text_area/data_editor의 이전 값이 새 설정을 덮어쓰지 않는다
EDITOR_KEYS = ['dispatch_editor', 'cost_editor', 'i3_editor', 'cons_editor',
               'exp_editor', 'i2_editor', 'ec_editor', 'ei3_editor', 'ei2_editor']


def reset_editor_widgets():
    for k in EDITOR_KEYS:
        st.session_state.pop(k, None)


st.set_page_config(page_title="부서별 월정산 구분 시스템", layout="wide")

st.markdown("""
<style>
html, body, [class*="css"], .stMarkdown, .stMetric, .stDataFrame, button {
    font-family: 'Malgun Gothic', 'Gulim', sans-serif !important;
    font-size: 12px !important;
}

/* 메트릭 카드 스타일 */
[data-testid="stMetricValue"] {
    font-size: 20px !important;
    font-weight: 700 !important;
}
[data-testid="stMetricLabel"] {
    font-size: 13px !important;
    font-weight: 600 !important;
}
[data-testid="stMetricDelta"] {
    font-size: 12px !important;
}

/* 데이터프레임 */
.stDataFrame div[data-testid="stTable"] {
    font-family: 'Malgun Gothic', 'Gulim', sans-serif !important;
}

/* 탭 스타일 */
.stTabs [data-baseweb="tab-list"] {
    gap: 4px;
}
.stTabs [data-baseweb="tab"] {
    padding: 8px 20px;
    font-weight: 600;
}

/* 다운로드 버튼 */
.stDownloadButton > button {
    border: 2px solid #4f46e5 !important;
    color: #4f46e5 !important;
    font-weight: 600 !important;
    transition: all 0.2s !important;
}
.stDownloadButton > button:hover {
    background: #4f46e5 !important;
    color: white !important;
}
</style>
""", unsafe_allow_html=True)


def load_config(config_file):
    config = {}
    try:
        file_ext = config_file.name.split('.')[-1].lower()
        engine = 'openpyxl' if file_ext in ['xlsx', 'xlsm'] else 'xlrd'

        df_import2 = pd.read_excel(config_file, sheet_name='설정_수입2팀', header=2, engine=engine)
        config['import2'] = df_import2[['받는자', '업무', '실화주']].fillna('').values.tolist()

        xls = pd.ExcelFile(config_file, engine=engine)
        df_export_raw = pd.read_excel(xls, sheet_name='설정_수출팀', header=None)

        mask_export_header = df_export_raw.astype(str).apply(lambda x: x.str.contains('대상 업무|등록자', na=False)).any(axis=1)

        if mask_export_header.any():
            header_row = df_export_raw[mask_export_header].index[0]
            header_col = df_export_raw.iloc[header_row].astype(str).str.contains('대상 업무|등록자', na=False).values.argmax()

            df_after_header = df_export_raw.iloc[header_row+1:]
            mask_spec_header = df_after_header.astype(str).apply(lambda x: x.str.contains('특정 업체', na=False)).any(axis=1)

            if mask_spec_header.any():
                spec_header_row = df_after_header[mask_spec_header].index[0]
            else:
                spec_header_row = len(df_export_raw)

            df_items = df_export_raw.iloc[header_row+1:spec_header_row, header_col].dropna()
            config['export_tasks'] = [str(i).strip() for i in df_items if str(i).strip() and str(i).lower() != 'nan']
        else:
            config['export_tasks'] = ['수출', '갈음', '환급']
            spec_header_row = 0

        df_after_spec = df_export_raw.iloc[spec_header_row:]
        mask_receiver = df_after_spec.astype(str).apply(lambda x: x.str.contains('받는자', na=False)).any(axis=1)

        if mask_receiver.any():
            data_start_row = df_after_spec[mask_receiver].index[0] + 1
            df_spec = df_export_raw.iloc[data_start_row:, :3]
            config['export_companies'] = df_spec.fillna('').values.tolist()
        else:
            config['export_companies'] = []

        df_cost = pd.read_excel(config_file, sheet_name='설정_매출원가', header=2, engine=engine)
        config['cost_categories'] = df_cost['과목명'].dropna().tolist()

        df_import3 = pd.read_excel(config_file, sheet_name='설정_수입3팀', header=2, engine=engine)
        config['import3_receivers'] = df_import3['받는자'].dropna().tolist()

        df_consulting = pd.read_excel(config_file, sheet_name='설정_컨설팅', header=2, engine=engine)
        config['consulting_tasks'] = df_consulting['업무'].dropna().tolist()

        return config

    except Exception as e:
        st.error(f"설정 파일을 읽는 중 오류가 발생했습니다: {e}")
        return None


def process_data(file, config):
    try:
        file_ext = file.name.split('.')[-1].lower()
        engine = 'openpyxl' if file_ext in ['xlsx', 'xlsm'] else 'xlrd'
        df = pd.read_excel(file, sheet_name='원본', engine=engine)
    except Exception as e:
        st.error(f"파일을 열 수 없습니다: {e}")
        return None

    for col in ['업무', '받는자', '실화주', '과목명', '등록자']:
        if col not in df.columns:
            df[col] = ''
            st.warning(f"원본에 '{col}' 컬럼이 없어 빈 값으로 처리합니다.")
        df[col] = df[col].astype(str).str.strip()

    def clean_numeric(val):
        if pd.isna(val) or str(val).strip() == '':
            return 0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        # 회계식 음수 표기 (1,234) 지원
        is_paren_negative = s.startswith('(') and s.endswith(')')
        cleaned = re.sub(r'[^0-9.\-]', '', s)
        try:
            num = float(cleaned) if cleaned else 0
        except ValueError:
            return 0
        return -abs(num) if is_paren_negative else num

    if '공급가' in df.columns:
        df['공급가'] = df['공급가'].apply(clean_numeric)
    else:
        potential_cols = ['매출액', '금액', '결제금액', '합계']
        found_col = next((c for c in potential_cols if c in df.columns), None)
        if found_col:
            df['공급가'] = df[found_col].apply(clean_numeric)
            st.warning(f"'공급가' 컬럼이 없어 '{found_col}' 컬럼을 사용합니다.")
        else:
            df['공급가'] = 0
            st.error("매출액을 집계할 수 있는 컬럼(공급가 등)을 찾을 수 없습니다.")

    # --- Cascading Priority Filtering ---

    # 0순위: 파견비용 (B/L No에 키워드 포함)
    dispatch_keywords = [str(k).strip() for k in config.get('dispatch_keywords', ['외주용역비', '파견비용청구']) if str(k).strip()]
    if dispatch_keywords and 'B/L No' in df.columns:
        dispatch_pattern = '|'.join(re.escape(k) for k in dispatch_keywords)
        df_dispatch = df[df['B/L No'].astype(str).str.contains(dispatch_pattern, na=False)]
    else:
        if dispatch_keywords and 'B/L No' not in df.columns:
            st.warning("원본에 'B/L No' 컬럼이 없어 파견비용 분류를 건너뜁니다.")
        df_dispatch = pd.DataFrame(columns=df.columns)
    df_rem = df[~df.index.isin(df_dispatch.index)]

    # 1순위: 매출원가 (과목명 일치 시 무조건 매출원가)
    cost_of_sales_categories = [str(c).strip() for c in config.get('cost_categories', [])]
    df_cost_of_sales = df_rem[df_rem['과목명'].isin(cost_of_sales_categories)]
    df_rem = df_rem[~df_rem.index.isin(df_cost_of_sales.index)]

    # 1순위: 컨설팅
    consulting_tasks = [str(t).strip() for t in config.get('consulting_tasks', ['기타'])]
    df_consulting = df_rem[df_rem['업무'].isin(consulting_tasks)]
    df_rem = df_rem[~df_rem.index.isin(df_consulting.index)]

    # 2순위: 수입3팀
    import3_receivers = [str(r).strip() for r in config.get('import3_receivers', ['삼성바이오로직스주식회사'])]
    df_income_team3 = df_rem[df_rem['받는자'].isin(import3_receivers)]
    df_rem = df_rem[~df_rem.index.isin(df_income_team3.index)]

    # 3순위: 수입2팀
    shipper_conditions = config.get('import2', [])
    filtered_dfs_i2 = []
    for row in shipper_conditions:
        if len(row) < 3:
            continue
        r, b, s = str(row[0]).strip(), str(row[1]).strip(), str(row[2]).strip()
        if s:
            f = df_rem[(df_rem['받는자'] == r) & (df_rem['업무'] == b) & (df_rem['실화주'] == s)]
        else:
            f = df_rem[(df_rem['받는자'] == r) & (df_rem['업무'] == b)]
        filtered_dfs_i2.append(f)

    filtered_dfs_i2 = [d for d in filtered_dfs_i2 if not d.empty]
    if filtered_dfs_i2:
        df_income_team2 = pd.concat(filtered_dfs_i2)
        df_income_team2 = df_income_team2[~df_income_team2.index.duplicated(keep='first')]
    else:
        df_income_team2 = pd.DataFrame(columns=df.columns)
    df_rem = df_rem[~df_rem.index.isin(df_income_team2.index)]

    # 4순위: 수출팀
    export_companies = config.get('export_companies', [])
    export_tasks = config.get('export_tasks', ['수출', '갈음', '환급'])

    task_pattern = '|'.join(re.escape(str(t).strip()) for t in export_tasks if str(t).strip())
    if task_pattern:
        mask_export_task = df_rem['업무'].str.contains(task_pattern, na=False)
        df_export_by_task = df_rem[mask_export_task]
    else:
        df_export_by_task = pd.DataFrame(columns=df.columns)

    export_company_dfs = []
    for row in export_companies:
        if len(row) < 3:
            continue
        rcv = str(row[0]).strip()
        tasks = [str(t).strip() for t in row[1:] if str(t).strip()]
        if tasks:
            filtered = df_rem[(df_rem['받는자'] == rcv) & (df_rem['업무'].isin(tasks))]
            if not filtered.empty:
                export_company_dfs.append(filtered)

    export_parts = [d for d in [df_export_by_task] + export_company_dfs if not d.empty]
    if export_parts:
        df_export = pd.concat(export_parts)
        df_export = df_export[~df_export.index.duplicated(keep='first')]
    else:
        df_export = pd.DataFrame(columns=df.columns)
    df_rem = df_rem[~df_rem.index.isin(df_export.index)]

    # 4-1순위: 수출팀 내 등록자 기반 재분류
    export_to_i3 = config.get('export_to_import3_registrars', [])
    export_to_i2 = config.get('export_to_import2_registrars', [])

    if export_to_i3 and '등록자' in df_export.columns:
        mask_to_i3 = df_export['등록자'].isin(export_to_i3)
        if mask_to_i3.any():
            parts = [d for d in (df_income_team3, df_export[mask_to_i3]) if not d.empty]
            df_income_team3 = pd.concat(parts)
            df_income_team3 = df_income_team3[~df_income_team3.index.duplicated(keep='first')]
            df_export = df_export[~mask_to_i3]

    if export_to_i2 and '등록자' in df_export.columns:
        mask_to_i2 = df_export['등록자'].isin(export_to_i2)
        if mask_to_i2.any():
            parts = [d for d in (df_income_team2, df_export[mask_to_i2]) if not d.empty]
            df_income_team2 = pd.concat(parts)
            df_income_team2 = df_income_team2[~df_income_team2.index.duplicated(keep='first')]
            df_export = df_export[~mask_to_i2]

    # 6순위: 수입1팀 (나머지)
    df_import_team1 = df_rem

    # --- 파견비용 누락 의심 감지 ---
    # B/L No에 '파견'/'용역'이 있는데 파견비용으로 분류되지 않은 행을 찾아 경고.
    # 담당자가 B/L 표기를 바꾸면(예: '파견비용청구-' → '파견비용-') 키워드가 빗나가
    # 매출원가 등으로 새어나가는 사고를 막기 위한 안전망.
    if 'B/L No' in df.columns:
        suspect_mask = df['B/L No'].astype(str).str.contains('파견|용역', na=False)
        missed = df[suspect_mask & ~df.index.isin(df_dispatch.index)]
        if not missed.empty:
            bl_list = ', '.join(missed['B/L No'].astype(str).head(10).tolist())
            st.warning(
                f"⚠️ 파견비용 누락 의심 {len(missed)}건 "
                f"({missed['공급가'].sum():,.0f}원) — B/L No에 '파견' 또는 '용역'이 있으나 "
                f"현재 키워드({', '.join(dispatch_keywords) or '없음'})와 맞지 않아 다른 부서로 분류되었습니다. "
                f"해당 B/L: {bl_list}"
            )

    return df, df_dispatch, df_consulting, df_income_team3, df_income_team2, df_export, df_cost_of_sales, df_import_team1


def create_summary(df, condition_column=None, condition_value=None, group_cols=None):
    """반환: (summary, fallback) — fallback=True면 필터 결과가 없어 팀 전체 기준으로 집계됨."""
    if group_cols is None:
        group_cols = ['받는자', '업무']

    if df.empty:
        return pd.DataFrame(columns=group_cols + ['매출액']), False

    if condition_column is None:
        filtered_data = df
    elif isinstance(condition_value, list):
        vals = [str(v).strip() for v in condition_value]
        filtered_data = df[df[condition_column].isin(vals)]
    else:
        val = str(condition_value).strip()
        filtered_data = df[df[condition_column] == val]

    fallback = False
    if filtered_data.empty and not df.empty:
        filtered_data = df
        fallback = True

    summary = filtered_data.groupby(group_cols, as_index=False)['공급가'].sum()
    summary.rename(columns={'공급가': '매출액'}, inplace=True)
    summary = summary.sort_values(by='매출액', ascending=False).reset_index(drop=True)
    return summary, fallback


def create_full_summary(df, group_cols=None):
    """팀 전체 데이터 요약 (필터 없이)."""
    if group_cols is None:
        group_cols = ['받는자', '업무']

    if df.empty:
        return pd.DataFrame(columns=group_cols + ['매출액'])

    summary = df.groupby(group_cols, as_index=False)['공급가'].sum()
    summary.rename(columns={'공급가': '매출액'}, inplace=True)
    summary = summary.sort_values(by='매출액', ascending=False).reset_index(drop=True)
    return summary


def create_validation_sheet(df, df_dispatch, df_consulting, df_income_team3, df_income_team2, df_export, df_cost_of_sales, df_import_team1):
    validation_data = {
        '구분': ['원본 데이터', '파견비용', '매출원가', '컨설팅', '수입3팀', '수출팀', '수입2팀', '수입1팀', '합계', '차이'],
        '공급가': [
            df['공급가'].sum(),
            df_dispatch['공급가'].sum(),
            df_cost_of_sales['공급가'].sum(),
            df_consulting['공급가'].sum(),
            df_income_team3['공급가'].sum(),
            df_export['공급가'].sum(),
            df_income_team2['공급가'].sum(),
            df_import_team1['공급가'].sum(),
            0, 0
        ],
        '건수': [
            len(df), len(df_dispatch), len(df_cost_of_sales), len(df_consulting),
            len(df_income_team3), len(df_export), len(df_income_team2), len(df_import_team1), 0, 0
        ]
    }

    df_validation = pd.DataFrame(validation_data)
    sum_supply = df_validation.iloc[1:8]['공급가'].sum()
    sum_count = df_validation.iloc[1:8]['건수'].sum()

    df_validation.loc[df_validation['구분'] == '합계', '공급가'] = sum_supply
    df_validation.loc[df_validation['구분'] == '합계', '건수'] = sum_count
    df_validation.loc[df_validation['구분'] == '차이', '공급가'] = df_validation.iloc[0]['공급가'] - sum_supply
    df_validation.loc[df_validation['구분'] == '차이', '건수'] = df_validation.iloc[0]['건수'] - sum_count

    return df_validation


def to_excel(df, df_dispatch, df_consulting, df_income_team3, df_income_team2, df_export, df_cost_of_sales, df_import_team1, config):
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='원본', index=False)

        # 합계는 1행에 별도 기록, 데이터 헤더는 2행부터 — 자동필터 범위에서 합계 제외
        total_sheets = {}  # sheet_name -> (합계, 공급가 컬럼 번호)

        def write_team_sheet(team_df, sheet_name):
            if team_df.empty:
                team_df.to_excel(writer, sheet_name=sheet_name, index=False)
                return
            team_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=1)
            total_sheets[sheet_name] = (
                team_df['공급가'].sum(),
                list(team_df.columns).index('공급가') + 1,
            )

        write_team_sheet(df_dispatch, '파견비용')
        write_team_sheet(df_consulting, '컨설팅')
        write_team_sheet(df_income_team3, '수입3팀')
        write_team_sheet(df_income_team2, '수입2팀')
        write_team_sheet(df_export, '수출')
        write_team_sheet(df_cost_of_sales, '매출원가')
        write_team_sheet(df_import_team1, '수입1팀')

        def write_summary_sheet(team_df, sheet_name, cond_col=None, cond_val=None, filter_desc=None):
            summary, fallback = create_summary(team_df, cond_col, cond_val)
            if filter_desc and not summary.empty:
                note = (f"※ {filter_desc} 항목 없음 — 팀 전체 기준 집계" if fallback
                        else f"※ {filter_desc} 기준 집계")
                note_row = pd.DataFrame({col: [''] for col in summary.columns})
                note_row[summary.columns[0]] = [note]
                summary = pd.concat([note_row, summary], ignore_index=True)
            summary.to_excel(writer, sheet_name=sheet_name, index=False)

        write_summary_sheet(df_consulting, '컨설팅 Summary', '업무', config.get('consulting_tasks', ['기타']), '컨설팅 업무')
        write_summary_sheet(df_income_team3, '수입3팀 Summary', '과목명', '통관수수료', '통관수수료')
        write_summary_sheet(df_income_team2, '수입2팀 Summary', '과목명', '통관수수료', '통관수수료')
        write_summary_sheet(df_export, '수출 Summary')
        write_summary_sheet(df_import_team1, '수입1팀 Summary', '과목명', '통관수수료', '통관수수료')

        df_validation = create_validation_sheet(df, df_dispatch, df_consulting, df_income_team3, df_income_team2, df_export, df_cost_of_sales, df_import_team1)
        df_validation.to_excel(writer, sheet_name='검증', index=False)

        accounting_format = '#,##0'
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        gulim_font = Font(name='굴림', size=9)
        gulim_bold = Font(name='굴림', size=9, bold=True)
        red_bold = Font(name='굴림', size=9, bold=True, color='FF0000')

        def cell_width(text):
            # 한글/한자 2칸 폭 근사
            return sum(2 if ord(ch) > 0x2E80 else 1 for ch in text)

        # 열너비는 상위 행 샘플로 충분 — 수천 행 전체를 재면 과도하게 느려짐
        WIDTH_SCAN_ROWS = 300

        for sheet_name in writer.sheets:
            ws = writer.sheets[sheet_name]
            has_total = sheet_name in total_sheets
            header_row = 2 if has_total else 1

            if has_total:
                total_val, supply_col = total_sheets[sheet_name]
                if supply_col != 1:
                    ws.cell(row=1, column=1, value='합계')
                ws.cell(row=1, column=supply_col, value=total_val)

            col_widths = {}
            for r_idx, row in enumerate(ws.iter_rows(), start=1):
                scan_width = r_idx <= WIDTH_SCAN_ROWS
                for cell in row:
                    cell.font = gulim_font
                    if scan_width and cell.value is not None:
                        text = str(cell.value)
                        if text:
                            w = cell_width(text)
                            if w > col_widths.get(cell.column, 0):
                                col_widths[cell.column] = w
            for col_idx, w in col_widths.items():
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max(w + 2, 8), 45)

            ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

            if has_total:
                for cell in ws[1]:
                    cell.font = gulim_bold
                    cell.fill = total_fill

            if sheet_name != '검증':
                for cell in ws[header_row]:
                    cell.fill = yellow_fill
                    cell.font = gulim_bold
                if ws.max_row > header_row:
                    last_col = ws.cell(row=header_row, column=ws.max_column).column_letter
                    ws.auto_filter.ref = f"A{header_row}:{last_col}{ws.max_row}"

            if sheet_name == '검증':
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == '차이':
                        try:
                            diff_val = float(row[1].value or 0)
                        except (TypeError, ValueError):
                            diff_val = 0
                        if abs(diff_val) >= 0.5:
                            for cell in row:
                                cell.font = red_bold

            for idx in range(1, ws.max_column + 1):
                if ws.cell(row=header_row, column=idx).value in ('공급가', '매출액'):
                    for r in range(1, ws.max_row + 1):
                        if r != header_row:
                            ws.cell(row=r, column=idx).number_format = accounting_format

    return output.getvalue()


def make_chart(df_summary, title=""):
    """차트 생성 공통 함수."""
    chart_df = df_summary.groupby('받는자', as_index=False)['매출액'].sum()
    top_10 = chart_df.sort_values('매출액', ascending=False).head(10).sort_values('매출액', ascending=True)
    top_10 = top_10.copy()
    top_10['매출액_텍스트'] = top_10['매출액'].apply(lambda x: f"{x:,.0f}")

    colors = ['#6366f1', '#818cf8', '#a5b4fc', '#c7d2fe', '#e0e7ff',
              '#7c3aed', '#8b5cf6', '#a78bfa', '#c4b5fd', '#ddd6fe']

    fig = px.bar(top_10, x='매출액', y='받는자', orientation='h',
                 text='매출액_텍스트',
                 labels={'매출액': '매출액(원)', '받는자': ''},
                 template='plotly_white')
    fig.update_traces(
        marker_color=colors[:len(top_10)],
        textposition='outside',
        cliponaxis=False,
        textfont=dict(size=11, color='#334155'),
        hovertemplate='<b>%{y}</b><br>매출액: %{text}원<extra></extra>'
    )
    fig.update_layout(
        margin=dict(l=10, r=130, t=10, b=10),
        height=380,
        plot_bgcolor='rgba(0,0,0,0)',
        paper_bgcolor='rgba(0,0,0,0)',
        xaxis=dict(showgrid=True, gridcolor='#f1f5f9', zeroline=False, showticklabels=False),
        yaxis=dict(showgrid=False, tickfont=dict(size=11)),
        showlegend=False
    )
    return fig


def display_dashboard(df_consulting, df_income_team3, df_income_team2, df_export, df_import_team1, config):
    st.subheader("3. 부서별 매출 현황")

    detailed_cols = ['받는자', '과목명', '업무']
    summaries = {
        "수입3팀": create_full_summary(df_income_team3, group_cols=detailed_cols),
        "수입2팀": create_full_summary(df_income_team2, group_cols=detailed_cols),
        "수출팀": create_full_summary(df_export, group_cols=detailed_cols),
        "수입1팀": create_full_summary(df_import_team1, group_cols=detailed_cols),
        "컨설팅": create_full_summary(df_consulting, group_cols=detailed_cols)
    }
    tabs = st.tabs(list(summaries.keys()))
    for tab, (team_name, summary_df) in zip(tabs, summaries.items()):
        with tab:
            if not summary_df.empty:
                total_revenue = summary_df['매출액'].sum()
                company_count = summary_df['받는자'].nunique()

                c1, c2, c3 = st.columns(3)
                c1.metric("총 매출액", f"{total_revenue:,.0f}원")
                c2.metric("업체 수", f"{company_count}개사")
                c3.metric("항목 수", f"{len(summary_df)}개")

                col1, col2 = st.columns([2, 1])
                with col1:
                    fig = make_chart(summary_df)
                    st.plotly_chart(fig, width="stretch")
                with col2:
                    st.markdown("**상세 데이터 (상위 20)**")
                    formatted_df = summary_df.head(20).copy()
                    st.dataframe(
                        formatted_df.style.format({"매출액": "{:,.0f}"}),
                        hide_index=True,
                        width="stretch"
                    )
            else:
                st.info("데이터가 없습니다.")
    return summaries


def generate_insights(summaries):
    if not summaries:
        return {"overall": "데이터가 없습니다.", "teams": {}}

    team_stats = []
    for team, df in summaries.items():
        if df.empty:
            team_stats.append({'team': team, 'revenue': 0, 'companies': 0, 'count': 0, 'top_company': '-', 'top_revenue': 0, 'avg_per_deal': 0})
            continue
        total = df['매출액'].sum()
        top = df.groupby('받는자')['매출액'].sum()
        top_company = top.idxmax() if not top.empty else '-'
        top_rev = top.max() if not top.empty else 0
        team_stats.append({
            'team': team, 'revenue': total,
            'companies': df['받는자'].nunique(), 'count': len(df),
            'top_company': top_company, 'top_revenue': top_rev,
            'avg_per_deal': total / len(df) if len(df) > 0 else 0
        })

    team_stats.sort(key=lambda x: x['revenue'], reverse=True)
    total_all = sum(s['revenue'] for s in team_stats)
    active_teams = [s for s in team_stats if s['revenue'] > 0]

    overall_parts = []
    overall_parts.append(f"이번 달 총 매출액은 {total_all:,.0f}원이며, {len(active_teams)}개 부서에서 매출이 발생했습니다.")

    if active_teams:
        top = active_teams[0]
        top_pct = (top['revenue'] / total_all * 100) if total_all else 0
        overall_parts.append(f"최대 매출 부서는 {top['team']}으로 전체의 {top_pct:.1f}%({top['revenue']:,.0f}원)를 차지합니다.")

    if len(active_teams) >= 3:
        top3_rev = sum(s['revenue'] for s in active_teams[:3])
        top3_pct = (top3_rev / total_all * 100) if total_all else 0
        top3_names = ', '.join(s['team'] for s in active_teams[:3])
        overall_parts.append(f"상위 3개 부서({top3_names})가 전체 매출의 {top3_pct:.1f}%를 점유하고 있습니다.")

    avg_per_deal_all = total_all / sum(s['count'] for s in active_teams) if active_teams else 0
    overall_parts.append(f"집계 항목당 평균 매출액은 {avg_per_deal_all:,.0f}원입니다.")

    team_insights = {}
    for s in team_stats:
        if s['revenue'] == 0:
            team_insights[s['team']] = "해당 기간 매출 없음."
            continue
        pct = (s['revenue'] / total_all * 100) if total_all else 0
        top_pct = (s['top_revenue'] / s['revenue'] * 100) if s['revenue'] else 0
        parts = []
        parts.append(f"전체 매출의 {pct:.1f}% 차지. {s['companies']}개 업체, {s['count']}개 집계 항목.")
        parts.append(f"최대 거래처: {s['top_company']} ({s['top_revenue']:,.0f}원, 부서 내 {top_pct:.1f}%).")
        parts.append(f"항목당 평균 {s['avg_per_deal']:,.0f}원.")
        if top_pct > 50:
            parts.append(f"⚠ {s['top_company']} 매출 집중도 높음 — 리스크 분산 검토 필요.")
        if s['companies'] == 1:
            parts.append("⚠ 단일 거래처 의존 구조.")
        team_insights[s['team']] = ' '.join(parts)

    return {"overall": ' '.join(overall_parts), "teams": team_insights}


def to_html(summaries, config, extra_summaries=None):
    # summaries: 매출 부서 5개 (인사이트·도넛·총합 기준)
    # extra_summaries: 파견비용·매출원가 등 참고 섹션 (총합·비중 계산에서 제외)
    insights = generate_insights(summaries)
    all_summaries = dict(summaries)
    if extra_summaries:
        all_summaries.update(extra_summaries)

    charts_json = {}
    tables_data = {}
    team_totals = {}

    for team, df in all_summaries.items():
        if not df.empty:
            chart_df = df.groupby('받는자', as_index=False)['매출액'].sum()
            top_10 = chart_df.sort_values('매출액', ascending=False).head(10).sort_values('매출액', ascending=True)
            top_10 = top_10.copy()
            top_10['매출액_텍스트'] = top_10['매출액'].apply(lambda x: f"{x:,.0f}")
            top_10['받는자_short'] = top_10['받는자'].apply(lambda x: x[:12] + '…' if len(str(x)) > 14 else x)

            colors = ['#6366f1', '#818cf8', '#a5b4fc', '#c7d2fe', '#e0e7ff',
                      '#7c3aed', '#8b5cf6', '#a78bfa', '#c4b5fd', '#ddd6fe']

            fig = px.bar(top_10, x='매출액', y='받는자_short', orientation='h',
                         text='매출액_텍스트',
                         labels={'매출액': '매출액(원)', '받는자_short': ''},
                         template='plotly_white',
                         custom_data=['받는자'])
            fig.update_traces(
                marker_color=colors[:len(top_10)],
                textposition='outside',
                cliponaxis=False,
                textfont=dict(size=11, color='#1e293b'),
                hovertemplate='<b>%{customdata[0]}</b><br>매출액: %{text}원<extra></extra>'
            )
            fig.update_layout(
                margin=dict(l=10, r=150, t=10, b=10),
                height=max(300, len(top_10) * 42),
                plot_bgcolor='rgba(0,0,0,0)',
                paper_bgcolor='rgba(0,0,0,0)',
                xaxis=dict(showgrid=True, gridcolor='#f1f5f9', zeroline=False, showticklabels=False),
                yaxis=dict(showgrid=False, tickfont=dict(size=11)),
                showlegend=False,
                bargap=0.3
            )
            charts_json[team] = fig.to_json()
            tables_data[team] = df.to_dict('records')
            team_totals[team] = {
                'revenue': df['매출액'].sum(),
                'companies': df['받는자'].nunique(),
                'count': len(df)
            }
        else:
            charts_json[team] = "{}"
            tables_data[team] = []
            team_totals[team] = {'revenue': 0, 'companies': 0, 'count': 0}

    # 총합 KPI (매출 부서만)
    revenue_teams = [t for t in summaries.keys()]
    total_kpi = {
        'revenue': sum(team_totals[t]['revenue'] for t in revenue_teams),
        'count': sum(team_totals[t]['count'] for t in revenue_teams),
        'teams': sum(1 for t in revenue_teams if team_totals[t]['revenue'] > 0),
    }

    # 부서별 매출 비중 도넛 (매출 부서만)
    donut_rows = [(t, team_totals[t]['revenue']) for t in revenue_teams if team_totals[t]['revenue'] > 0]
    if donut_rows:
        donut_df = pd.DataFrame(donut_rows, columns=['부서', '매출액'])
        donut_colors = ['#1e3a8a', '#4f46e5', '#7c3aed', '#0891b2', '#059669', '#d97706']
        fig_d = px.pie(donut_df, names='부서', values='매출액', hole=0.55)
        fig_d.update_traces(
            marker=dict(colors=donut_colors[:len(donut_df)]),
            textinfo='label+percent',
            textfont=dict(size=12),
            hovertemplate='<b>%{label}</b><br>매출액: %{value:,.0f}원<br>비중: %{percent}<extra></extra>'
        )
        fig_d.update_layout(
            margin=dict(l=10, r=10, t=10, b=10),
            height=340,
            showlegend=True,
            legend=dict(orientation='v', font=dict(size=12)),
            paper_bgcolor='rgba(0,0,0,0)'
        )
        donut_json = fig_d.to_json()
    else:
        donut_json = "{}"

    html_template = """
    <!DOCTYPE html>
    <html lang="ko">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>월정산 분석 보고서</title>
        <script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
        <script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
        <link rel="stylesheet" href="https://cdn.datatables.net/1.13.6/css/jquery.dataTables.min.css">
        <script src="https://cdn.datatables.net/1.13.6/js/jquery.dataTables.min.js"></script>
        <link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@300;400;500;700;900&display=swap" rel="stylesheet">
        <style>
            :root {
                --primary: #1e3a8a;
                --primary-light: #3b82f6;
                --accent: #4f46e5;
                --accent-soft: #6366f1;
                --accent-bg: #eef2ff;
                --bg: #f8fafc;
                --card: #ffffff;
                --text: #1e293b;
                --text-sub: #475569;
                --text-muted: #94a3b8;
                --border: #e2e8f0;
                --border-light: #f1f5f9;
                --green: #059669;
                --green-bg: #ecfdf5;
                --amber: #d97706;
                --amber-bg: #fffbeb;
                --shadow-sm: 0 1px 3px rgba(0,0,0,0.06);
                --shadow: 0 4px 16px rgba(30,58,138,0.06);
                --shadow-lg: 0 8px 32px rgba(30,58,138,0.10);
                --radius: 12px;
                --radius-sm: 8px;
            }

            * { box-sizing: border-box; margin: 0; padding: 0; }

            body {
                font-family: 'Noto Sans KR', -apple-system, BlinkMacSystemFont, sans-serif;
                line-height: 1.6;
                color: var(--text);
                background: var(--bg);
                -webkit-font-smoothing: antialiased;
            }

            .page { max-width: 1100px; margin: 0 auto; padding: 32px 24px; }

            /* ── 헤더 ── */
            .hero {
                background: linear-gradient(135deg, #1e3a8a 0%, #4338ca 50%, #6366f1 100%);
                color: white;
                border-radius: 20px;
                padding: 52px 48px 44px;
                margin-bottom: 32px;
                box-shadow: var(--shadow-lg);
                position: relative;
                overflow: hidden;
            }
            .hero::after {
                content: '';
                position: absolute;
                top: -80px; right: -80px;
                width: 320px; height: 320px;
                background: radial-gradient(circle, rgba(255,255,255,0.08) 0%, transparent 70%);
                border-radius: 50%;
            }
            .hero::before {
                content: '';
                position: absolute;
                bottom: -60px; left: 30%;
                width: 240px; height: 240px;
                background: radial-gradient(circle, rgba(255,255,255,0.04) 0%, transparent 70%);
                border-radius: 50%;
            }
            .hero h1 {
                font-size: 32px;
                font-weight: 900;
                letter-spacing: -1px;
                margin-bottom: 6px;
                position: relative;
            }
            .hero .tagline {
                font-size: 15px;
                font-weight: 300;
                opacity: 0.85;
                position: relative;
            }
            .hero .stamp {
                position: absolute;
                top: 28px; right: 36px;
                font-size: 12px;
                background: rgba(255,255,255,0.15);
                backdrop-filter: blur(4px);
                padding: 6px 16px;
                border-radius: 24px;
                font-weight: 400;
                letter-spacing: 0.3px;
            }

            /* ── KPI 행 ── */
            .kpi-row {
                display: grid;
                grid-template-columns: repeat(auto-fit, minmax(180px, 1fr));
                gap: 16px;
                margin-bottom: 32px;
            }
            .kpi {
                background: var(--card);
                border-radius: var(--radius);
                padding: 22px 24px;
                box-shadow: var(--shadow-sm);
                border: 1px solid var(--border-light);
                text-align: center;
            }
            .kpi .label { font-size: 12px; color: var(--text-muted); font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; margin-bottom: 6px; }
            .kpi .value { font-size: 22px; font-weight: 800; color: var(--primary); font-variant-numeric: tabular-nums; }
            .kpi .sub { font-size: 11px; color: var(--text-muted); margin-top: 4px; }
            .kpi.total {
                border: 2px solid var(--accent);
                background: linear-gradient(135deg, var(--accent-bg) 0%, #ffffff 100%);
            }
            .kpi.total .value { color: var(--accent); }

            /* ── 도넛 차트 카드 ── */
            .donut-card {
                background: var(--card);
                border-radius: var(--radius);
                padding: 20px;
                box-shadow: var(--shadow-sm);
                border: 1px solid var(--border-light);
            }

            /* ── 인사이트 카드 ── */
            .insight-section {
                margin-bottom: 32px;
            }
            .insight-section-title {
                font-size: 16px;
                font-weight: 700;
                color: var(--primary);
                margin-bottom: 16px;
                display: flex;
                align-items: center;
                gap: 8px;
            }
            .insight {
                background: var(--card);
                border-radius: var(--radius);
                padding: 22px 26px;
                margin-bottom: 14px;
                box-shadow: var(--shadow-sm);
                border: 1px solid var(--border-light);
                border-left: 4px solid var(--accent);
                display: flex;
                gap: 14px;
                align-items: flex-start;
            }
            .insight.overall {
                border-left-color: var(--green);
                background: linear-gradient(135deg, #f0fdf4 0%, #ffffff 100%);
            }
            .insight.warning { border-left-color: var(--amber); }
            .insight-icon {
                width: 36px; height: 36px;
                background: var(--accent-bg);
                border-radius: 10px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 16px;
                flex-shrink: 0;
            }
            .insight.overall .insight-icon { background: var(--green-bg); }
            .insight .team-tag {
                display: inline-block;
                background: var(--accent-bg);
                color: var(--accent);
                font-size: 11px;
                font-weight: 700;
                padding: 2px 10px;
                border-radius: 12px;
                margin-bottom: 6px;
            }
            .insight p { color: var(--text-sub); font-size: 13px; line-height: 1.8; }

            /* ── 팀 카드 ── */
            .team-card {
                background: var(--card);
                border-radius: var(--radius);
                margin-bottom: 28px;
                box-shadow: var(--shadow);
                overflow: hidden;
                border: 1px solid var(--border-light);
                page-break-inside: avoid;
            }

            .team-head {
                background: linear-gradient(135deg, var(--primary) 0%, var(--accent-soft) 100%);
                color: white;
                padding: 18px 28px;
                display: flex;
                align-items: center;
                gap: 14px;
            }
            .team-head .icon {
                width: 38px; height: 38px;
                background: rgba(255,255,255,0.18);
                border-radius: 10px;
                display: flex;
                align-items: center;
                justify-content: center;
                font-size: 16px;
                font-weight: 800;
            }
            .team-head .title { font-size: 18px; font-weight: 700; }

            .team-stats {
                display: grid;
                grid-template-columns: repeat(3, 1fr);
                border-bottom: 1px solid var(--border-light);
            }
            .team-stat {
                padding: 16px 24px;
                text-align: center;
                border-right: 1px solid var(--border-light);
            }
            .team-stat:last-child { border-right: none; }
            .team-stat .stat-label { font-size: 11px; color: var(--text-muted); font-weight: 500; text-transform: uppercase; letter-spacing: 0.5px; }
            .team-stat .stat-value { font-size: 18px; font-weight: 700; color: var(--accent); margin-top: 4px; font-variant-numeric: tabular-nums; }

            .team-body { padding: 28px; }

            .tbl-label {
                font-size: 14px;
                font-weight: 700;
                color: var(--text);
                margin-bottom: 14px;
                padding-bottom: 8px;
                border-bottom: 2px solid var(--border);
            }

            /* ── DataTable ── */
            table.dataTable { border-collapse: collapse !important; width: 100% !important; font-size: 13px !important; }
            table.dataTable thead th {
                background: var(--accent-bg) !important;
                color: var(--primary) !important;
                font-weight: 600 !important;
                border-bottom: 2px solid var(--accent) !important;
                padding: 11px 14px !important;
                font-size: 12px !important;
                letter-spacing: 0.3px;
            }
            table.dataTable tbody td {
                padding: 9px 14px !important;
                border-bottom: 1px solid var(--border-light) !important;
            }
            table.dataTable tbody tr:hover { background: #f8fafc !important; }
            table.dataTable tbody tr:nth-child(even) { background: #fafcfe; }

            td.amt {
                font-variant-numeric: tabular-nums;
                font-weight: 600;
                color: var(--primary);
                text-align: right !important;
            }

            .dataTables_wrapper .dataTables_filter input {
                border: 1px solid var(--border); border-radius: 6px;
                padding: 5px 12px; font-size: 13px; outline: none;
            }
            .dataTables_wrapper .dataTables_filter input:focus {
                border-color: var(--accent);
                box-shadow: 0 0 0 3px rgba(99,102,241,0.12);
            }
            .dataTables_wrapper .dataTables_length select { border: 1px solid var(--border); border-radius: 6px; padding: 3px 8px; }
            .dataTables_wrapper .dataTables_paginate .paginate_button.current {
                background: var(--accent) !important; color: white !important;
                border: none !important; border-radius: 6px;
            }
            .dataTables_wrapper .dataTables_paginate .paginate_button:hover {
                background: var(--accent-bg) !important; color: var(--accent) !important; border: none !important;
            }
            .dataTables_wrapper .dataTables_info { font-size: 11px; color: var(--text-muted); }

            /* ── 푸터 ── */
            .foot { text-align: center; padding: 36px 0 24px; color: var(--text-muted); font-size: 12px; }
            .foot .brand { font-weight: 700; color: var(--primary); }

            /* ── 테이블 래퍼 (가로 스크롤) ── */
            .table-wrap {
                overflow-x: auto;
                -webkit-overflow-scrolling: touch;
                margin: 0 -4px;
                padding: 0 4px;
            }

            /* ── 차트 래퍼 ── */
            .chart-wrap {
                overflow: visible;
                margin-bottom: 24px;
                min-height: 200px;
            }

            /* ── 반응형: 태블릿 ── */
            @media (max-width: 768px) {
                .page { padding: 16px 10px; }
                .hero { padding: 32px 20px 28px; border-radius: 14px; }
                .hero h1 { font-size: 22px; }
                .hero .tagline { font-size: 13px; }
                .hero .stamp { position: static; display: inline-block; margin-top: 12px; font-size: 11px; }
                .kpi-row { grid-template-columns: repeat(2, 1fr); gap: 10px; }
                .kpi { padding: 16px 14px; }
                .kpi .value { font-size: 18px; }
                .kpi .label { font-size: 11px; }
                .kpi .sub { font-size: 10px; }
                .team-stats { grid-template-columns: 1fr; }
                .team-stat { border-right: none; border-bottom: 1px solid var(--border-light); padding: 12px 18px; }
                .team-stat .stat-value { font-size: 16px; }
                .team-body { padding: 18px 14px; }
                .team-head { padding: 14px 18px; }
                .team-head .title { font-size: 16px; }
                .insight { flex-direction: column; gap: 8px; padding: 18px 16px; }
                .insight p { font-size: 12px; line-height: 1.7; }
                .insight-section-title { font-size: 14px; }
                table.dataTable { font-size: 12px !important; }
                table.dataTable thead th { padding: 8px 10px !important; font-size: 11px !important; white-space: nowrap; }
                table.dataTable tbody td { padding: 7px 10px !important; }
                .dataTables_wrapper .dataTables_filter input { width: 120px; font-size: 12px; }
                .dataTables_wrapper .dataTables_length select { font-size: 12px; }
                .tbl-label { font-size: 13px; }
            }

            /* ── 반응형: 모바일 ── */
            @media (max-width: 480px) {
                .page { padding: 12px 8px; }
                .hero { padding: 28px 16px 24px; border-radius: 12px; }
                .hero h1 { font-size: 19px; letter-spacing: -0.5px; }
                .hero .tagline { font-size: 12px; }
                .kpi-row { grid-template-columns: repeat(2, 1fr); gap: 8px; }
                .kpi { padding: 14px 10px; }
                .kpi .value { font-size: 15px; }
                .kpi .label { font-size: 10px; }
                .kpi .sub { font-size: 9px; }
                .team-card { margin-bottom: 20px; border-radius: 10px; }
                .team-head { padding: 12px 14px; gap: 10px; }
                .team-head .icon { width: 32px; height: 32px; font-size: 13px; }
                .team-head .title { font-size: 15px; }
                .team-stats { grid-template-columns: repeat(3, 1fr); }
                .team-stat { padding: 10px 8px; }
                .team-stat .stat-label { font-size: 9px; }
                .team-stat .stat-value { font-size: 14px; }
                .team-body { padding: 14px 10px; }
                .insight { padding: 14px 12px; margin-bottom: 10px; }
                .insight p { font-size: 11px; line-height: 1.6; }
                .insight-icon { width: 30px; height: 30px; font-size: 14px; }
                .insight .team-tag { font-size: 10px; padding: 1px 8px; }
                .insight-section { margin-bottom: 24px; }
                .insight-section-title { font-size: 13px; margin-bottom: 10px; }
                table.dataTable { font-size: 11px !important; }
                table.dataTable thead th { padding: 6px 6px !important; font-size: 10px !important; }
                table.dataTable tbody td { padding: 6px 6px !important; max-width: 120px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }
                td.amt { font-size: 11px; }
                .dataTables_wrapper .dataTables_filter,
                .dataTables_wrapper .dataTables_length { font-size: 11px; margin-bottom: 8px; }
                .dataTables_wrapper .dataTables_filter input { width: 100px; font-size: 11px; padding: 4px 8px; }
                .dataTables_wrapper .dataTables_info { font-size: 10px; }
                .dataTables_wrapper .dataTables_paginate .paginate_button { padding: 4px 8px !important; font-size: 11px; }
                .tbl-label { font-size: 12px; margin-bottom: 10px; }
                .foot { padding: 24px 0 16px; font-size: 11px; }
            }

            /* ── 반응형: 초소형 모바일 ── */
            @media (max-width: 360px) {
                .hero h1 { font-size: 17px; }
                .kpi-row { grid-template-columns: 1fr 1fr; gap: 6px; }
                .kpi .value { font-size: 14px; }
                .team-stats { grid-template-columns: 1fr; }
                .team-stat { border-right: none; }
                table.dataTable tbody td { max-width: 90px; font-size: 10px !important; }
            }

            /* ── 인쇄 ── */
            @media print {
                body { background: white; }
                .page { max-width: 100%; padding: 0; }
                .hero, .team-card, .insight, .kpi { box-shadow: none !important; break-inside: avoid; }
                .hero, .team-head { -webkit-print-color-adjust: exact; print-color-adjust: exact; }
                .dataTables_filter, .dataTables_length, .dataTables_paginate, .dataTables_info { display: none; }
                .page-break { page-break-before: always; }
            }
        </style>
    </head>
    <body>
    <div class="page">
        <div class="hero">
            <div class="stamp">{{ current_date }}</div>
            <h1>월정산 분석 보고서</h1>
            <div class="tagline">부서별 매출 현황 및 상세 내역 &middot; 관세법인 우신</div>
        </div>

        <div class="kpi-row">
            <div class="kpi total">
                <div class="label">전체 매출</div>
                <div class="value">{{ "{:,.0f}".format(total_kpi.revenue) }}</div>
                <div class="sub">{{ total_kpi.teams }}개 부서 &middot; {{ total_kpi.count }}개 항목</div>
            </div>
            {% for team, info in team_totals.items() %}
            <div class="kpi">
                <div class="label">{{ team }}</div>
                <div class="value">{{ "{:,.0f}".format(info.revenue) }}</div>
                <div class="sub">{{ info.companies }}개사 &middot; {{ info.count }}개 항목</div>
            </div>
            {% endfor %}
        </div>

        <div class="insight-section">
            <div class="insight-section-title">📊 종합 인사이트</div>
            <div class="insight overall">
                <div class="insight-icon">📈</div>
                <div><p>{{ insights.overall }}</p></div>
            </div>
        </div>

        <div class="insight-section">
            <div class="insight-section-title">🍩 부서별 매출 비중</div>
            <div class="donut-card">
                <div id="donut-chart"></div>
            </div>
        </div>

        <div class="insight-section">
            <div class="insight-section-title">🏢 부서별 인사이트</div>
            {% for team, text in insights.teams.items() %}
            <div class="insight {% if '⚠' in text %}warning{% endif %}">
                <div class="insight-icon">{{ icons_map.get(team, '📋') }}</div>
                <div>
                    <div class="team-tag">{{ team }}</div>
                    <p>{{ text }}</p>
                </div>
            </div>
            {% endfor %}
        </div>

        {% set icons = {'수입3팀':'3', '수입2팀':'2', '수출팀':'Ex', '수입1팀':'1', '컨설팅':'C', '파견비용':'파', '매출원가':'원'} %}
        {% for team, chart_json in charts_json.items() %}
        <div class="team-card {% if loop.index > 1 %}page-break{% endif %}">
            <div class="team-head">
                <div class="icon">{{ icons.get(team, '#') }}</div>
                <div class="title">{{ team }}</div>
            </div>
            <div class="team-stats">
                <div class="team-stat">
                    <div class="stat-label">매출액</div>
                    <div class="stat-value">{{ "{:,.0f}".format(team_totals[team].revenue) }}</div>
                </div>
                <div class="team-stat">
                    <div class="stat-label">업체 수</div>
                    <div class="stat-value">{{ team_totals[team].companies }}</div>
                </div>
                <div class="team-stat">
                    <div class="stat-label">항목 수</div>
                    <div class="stat-value">{{ team_totals[team].count }}</div>
                </div>
            </div>
            <div class="team-body">
                {% if insights.teams.get(team) %}
                <div style="background:var(--accent-bg);border-radius:var(--radius-sm);padding:14px 18px;margin-bottom:20px;font-size:13px;color:var(--text-sub);line-height:1.7;">
                    💡 {{ insights.teams[team] }}
                </div>
                {% endif %}
                <div class="chart-wrap" id="chart-{{ loop.index }}"></div>
                <div class="tbl-label">상세 내역</div>
                <div class="table-wrap">
                <table id="table-{{ loop.index }}" class="display" style="width:100%">
                    <thead>
                        <tr>
                            <th>받는자</th>
                            <th>과목명</th>
                            <th>업무</th>
                            <th style="text-align:right">매출액(원)</th>
                        </tr>
                    </thead>
                    <tbody>
                        {% for row in tables_data[team] %}
                        <tr>
                            <td>{{ row['받는자'] }}</td>
                            <td>{{ row['과목명'] }}</td>
                            <td>{{ row['업무'] }}</td>
                            <td class="amt">{{ "{:,.0f}".format(row['매출액']) }}</td>
                        </tr>
                        {% endfor %}
                    </tbody>
                </table>
                </div>
            </div>
        </div>
        {% endfor %}

        <div class="foot">
            <p><span class="brand">관세법인 우신</span> &middot; 부서별 월정산 구분 시스템</p>
            <p>&copy; {{ current_year }} All rights reserved.</p>
        </div>
    </div>

    <script>
    $(document).ready(function() {
        (function() {
            var donut = {{ donut_json | safe }};
            if (donut && donut.data) {
                donut.layout.autosize = true;
                donut.layout.font = {family: 'Noto Sans KR, sans-serif'};
                Plotly.newPlot('donut-chart', donut.data, donut.layout, {responsive: true, displayModeBar: false});
            }
        })();

        {% for team, chart_json in charts_json.items() %}
        (function() {
            var d = {{ chart_json | safe }};
            if (d && d.data) {
                d.layout.autosize = true;
                d.layout.font = {family: 'Noto Sans KR, sans-serif'};
                Plotly.newPlot('chart-{{ loop.index }}', d.data, d.layout, {responsive: true, displayModeBar: false});
            }
        })();
        {% endfor %}

        {% for team in tables_data.keys() %}
        $('#table-{{ loop.index }}').DataTable({
            "language": { "url": "https://cdn.datatables.net/plug-ins/1.13.6/i18n/ko.json" },
            "pageLength": 15,
            "order": [[ 3, "desc" ]],
            "dom": '<"top"fl>rt<"bottom"ip>',
            "columnDefs": [{ "className": "amt", "targets": 3 }]
        });
        {% endfor %}
    });
    </script>
    </body>
    </html>
    """

    icons_map = {'수입3팀':'📦', '수입2팀':'📦', '수출팀':'🚢', '수입1팀':'📦', '컨설팅':'💼', '파견비용':'🚚', '매출원가':'🧾'}

    template = Template(html_template, autoescape=True)
    return template.render(
        current_date=datetime.now().strftime("%Y-%m-%d %H:%M"),
        current_year=datetime.now().year,
        insights=insights,
        icons_map=icons_map,
        charts_json=charts_json,
        tables_data=tables_data,
        team_totals=team_totals,
        total_kpi=total_kpi,
        donut_json=donut_json
    )


# ═══════════════════════════════════════
# 사이드바: 설정 관리
# ═══════════════════════════════════════

if 'config' not in st.session_state:
    st.session_state.config = load_json_config()

with st.sidebar:
    st.header("설정 관리")

    st.markdown("""
    **분류 우선순위**
    | 순위 | 부서 | 기준 |
    |:---:|------|------|
    | 0 | **파견비용** | B/L No 키워드 |
    | 1 | **매출원가** | 과목명 일치 |
    | 2 | 컨설팅 | 업무 일치 |
    | 3 | 수입3팀 | 받는자 일치 |
    | 4 | 수입2팀 | 받는자+업무+실화주 |
    | 5 | 수출팀 | 업무 or 특정업체 |
    | 5-1 | 수출→수입3 | 등록자 재분류 |
    | 5-2 | 수출→수입2 | 등록자 재분류 |
    | 6 | 수입1팀 | 나머지 전부 |
    """)
    st.divider()

    config_tab1, config_tab2 = st.tabs(["직접 편집", "엑셀 가져오기"])

    with config_tab1:
        st.caption("설정을 직접 편집하고 저장합니다.")

        with st.expander("파견비용 B/L 키워드", expanded=False):
            dispatch_text = st.text_area(
                "키워드 (줄바꿈으로 구분)",
                value="\n".join(st.session_state.config.get('dispatch_keywords', ['외주용역비', '파견비용청구'])),
                height=80,
                key="dispatch_editor"
            )
            st.session_state.config['dispatch_keywords'] = [
                x.strip() for x in dispatch_text.split('\n') if x.strip()
            ]

        with st.expander("매출원가 과목명", expanded=False):
            cost_text = st.text_area(
                "과목명 (줄바꿈으로 구분)",
                value="\n".join(st.session_state.config.get('cost_categories', [])),
                height=200,
                key="cost_editor"
            )
            st.session_state.config['cost_categories'] = [
                x.strip() for x in cost_text.split('\n') if x.strip()
            ]

        with st.expander("수입3팀 받는자", expanded=False):
            i3_text = st.text_area(
                "받는자 (줄바꿈으로 구분)",
                value="\n".join(st.session_state.config.get('import3_receivers', [])),
                height=100,
                key="i3_editor"
            )
            st.session_state.config['import3_receivers'] = [
                x.strip() for x in i3_text.split('\n') if x.strip()
            ]

        with st.expander("컨설팅 업무", expanded=False):
            cons_text = st.text_area(
                "업무 (줄바꿈으로 구분)",
                value="\n".join(st.session_state.config.get('consulting_tasks', [])),
                height=80,
                key="cons_editor"
            )
            st.session_state.config['consulting_tasks'] = [
                x.strip() for x in cons_text.split('\n') if x.strip()
            ]

        with st.expander("수출팀 대상 업무", expanded=False):
            exp_text = st.text_area(
                "업무 (줄바꿈으로 구분)",
                value="\n".join(st.session_state.config.get('export_tasks', [])),
                height=80,
                key="exp_editor"
            )
            st.session_state.config['export_tasks'] = [
                x.strip() for x in exp_text.split('\n') if x.strip()
            ]

        with st.expander("수입2팀 업체 조건", expanded=False):
            i2_data = st.session_state.config.get('import2', [])
            df_i2 = pd.DataFrame(i2_data, columns=['받는자', '업무', '실화주'])
            edited_i2 = st.data_editor(
                df_i2,
                num_rows="dynamic",
                width="stretch",
                key="i2_editor"
            )
            st.session_state.config['import2'] = edited_i2.fillna('').values.tolist()

        with st.expander("수출팀 특정 업체", expanded=False):
            ec_data = st.session_state.config.get('export_companies', [])
            if not ec_data:
                ec_data = [["", "", ""]]
            df_ec = pd.DataFrame(ec_data, columns=['받는자', '업무1', '업무2'])
            edited_ec = st.data_editor(
                df_ec,
                num_rows="dynamic",
                width="stretch",
                key="ec_editor"
            )
            result_ec = edited_ec.fillna('').astype(str).values.tolist()
            st.session_state.config['export_companies'] = [
                r for r in result_ec if r[0].strip()
            ]

        with st.expander("수출팀→수입3팀 등록자", expanded=False):
            ei3_text = st.text_area(
                "등록자 (줄바꿈으로 구분)",
                value="\n".join(st.session_state.config.get('export_to_import3_registrars', [])),
                height=80,
                key="ei3_editor"
            )
            st.session_state.config['export_to_import3_registrars'] = [
                x.strip() for x in ei3_text.split('\n') if x.strip()
            ]

        with st.expander("수출팀→수입2팀 등록자", expanded=False):
            ei2_text = st.text_area(
                "등록자 (줄바꿈으로 구분)",
                value="\n".join(st.session_state.config.get('export_to_import2_registrars', [])),
                height=80,
                key="ei2_editor"
            )
            st.session_state.config['export_to_import2_registrars'] = [
                x.strip() for x in ei2_text.split('\n') if x.strip()
            ]

        if st.button("설정 저장", type="primary", width="stretch"):
            save_json_config(st.session_state.config)
            st.success("저장 완료!")

        st.caption("⚠️ 클라우드 배포에서는 앱이 재시작되면 저장한 설정이 초기화될 수 있습니다. 아래 버튼으로 JSON 백업을 받아두세요.")
        st.download_button(
            "설정 JSON 백업 다운로드",
            data=json.dumps(st.session_state.config, ensure_ascii=False, indent=2),
            file_name=f"월정산_설정백업_{datetime.now().strftime('%Y%m%d')}.json",
            mime="application/json",
            width="stretch"
        )

        st.divider()
        if st.button("기본값으로 초기화", width="stretch"):
            st.session_state.config = DEFAULT_CONFIG.copy()
            save_json_config(st.session_state.config)
            reset_editor_widgets()
            st.success("초기화 완료!")
            st.rerun()

    with config_tab2:
        st.caption("기존 엑셀 설정 파일에서 가져옵니다.")
        uploaded_config = st.file_uploader(
            "설정 파일 (.xlsx)",
            type=['xlsx', 'xls', 'xlsm'],
            key="config_uploader"
        )
        if uploaded_config:
            # 같은 파일 재처리로 인한 무한 rerun 방지
            file_id = f"{uploaded_config.name}-{uploaded_config.size}"
            if st.session_state.get('last_imported_config') != file_id:
                imported = load_config(uploaded_config)
                if imported:
                    # 엑셀에 없는 키(파견 키워드, 등록자 재분류 등)는 기존 설정 유지
                    merged = {**DEFAULT_CONFIG, **st.session_state.config, **imported}
                    st.session_state.config = merged
                    save_json_config(merged)
                    st.session_state['last_imported_config'] = file_id
                    reset_editor_widgets()
                    st.success("엑셀 설정을 가져와 저장했습니다!")
                    st.rerun()

        st.divider()
        st.caption("JSON 백업 파일로 설정을 복원합니다.")
        uploaded_json = st.file_uploader(
            "설정 JSON 복원 (.json)",
            type=['json'],
            key="json_uploader"
        )
        if uploaded_json:
            json_id = f"{uploaded_json.name}-{uploaded_json.size}"
            if st.session_state.get('last_restored_json') != json_id:
                try:
                    restored = json.load(uploaded_json)
                    if not isinstance(restored, dict):
                        raise ValueError("JSON 최상위가 객체(dict)가 아닙니다.")
                    merged = {**DEFAULT_CONFIG, **restored}
                    st.session_state.config = merged
                    save_json_config(merged)
                    st.session_state['last_restored_json'] = json_id
                    reset_editor_widgets()
                    st.success("JSON 설정을 복원했습니다!")
                    st.rerun()
                except Exception as e:
                    st.error(f"JSON 파일을 읽을 수 없습니다: {e}")

    st.divider()
    with st.expander("현재 설정 (JSON)"):
        st.json(st.session_state.config)


# ═══════════════════════════════════════
# 메인 UI
# ═══════════════════════════════════════

config = st.session_state.config

st.title("부서별 월정산 구분 시스템")
st.caption("관세법인 우신 · 본사 월정산 자동 분류 도구")
st.markdown("---")

with st.expander("이 프로그램은 무엇인가요? (처음 사용하시는 분은 클릭)", expanded=False):
    st.markdown("""
    ### 왜 만들어졌나요?

    매달 본사 월정산 작업 시, **수백~수천 건의 매출 데이터**를 부서별로 수작업 분류하면
    시간이 오래 걸리고 실수가 발생할 수 있습니다.

    이 시스템은 **설정한 규칙에 따라 자동으로 부서별 분류**해주어,
    정산 업무 시간을 대폭 줄이고 정확도를 높여줍니다.

    ---

    ### 사용 방법 (3단계)

    **1단계: 원본 파일 업로드**
    - 아래 업로드 버튼으로 월정산 원본 엑셀 파일(.xlsx)을 올립니다
    - 파일에 **'원본'** 시트가 있어야 합니다

    **2단계: "데이터 처리 시작" 클릭**
    - 자동으로 7개 부서로 분류됩니다:
      파견비용 → 매출원가 → 컨설팅 → 수입3팀 → 수입2팀 → 수출팀 → 수입1팀(나머지)

    **3단계: 결과 다운로드**
    - **엑셀 파일**: 부서별 시트로 나뉜 결과물
    - **HTML 보고서**: 차트와 인사이트가 포함된 시각 보고서

    ---

    ### 분류 규칙은 어떻게 바꾸나요?

    왼쪽 **사이드바** > **설정 관리**에서:
    - **직접 편집**: 거래처, 과목명, 키워드 등을 직접 수정
    - **엑셀 가져오기**: 기존 설정 엑셀 파일로 한번에 반영

    설정 변경 후 **"설정 저장"** 버튼을 꼭 눌러주세요.
    """)

if is_config_empty(config):
    st.warning("⚠️ 설정이 비어있습니다. 사이드바에서 직접 입력하거나 엑셀 설정 파일을 가져와주세요.")

uploaded_file = st.file_uploader(
    "원본 데이터 파일 업로드 (.xlsx / .xls)",
    type=['xlsx', 'xls', 'xlsm'],
    key="data_uploader"
)

if uploaded_file:
    if st.button("데이터 처리 시작", type="primary", width="stretch"):
        with st.spinner('처리 중...'):
            result = process_data(uploaded_file, config)
            if result is not None:
                st.session_state.result = result
                st.session_state.result_file_name = uploaded_file.name

    # 결과를 session_state에 보관 — 다운로드 버튼 클릭(rerun) 후에도 화면 유지
    if st.session_state.get('result') is not None and st.session_state.get('result_file_name') == uploaded_file.name:
        df, df_dispatch, df_consulting, df_income_team3, df_income_team2, df_export, df_cost_of_sales, df_import_team1 = st.session_state.result

        st.markdown("---")
        st.subheader("처리 결과")

        def fmt_money(val):
            return f"{int(val):,}"

        cols = st.columns(7)
        teams_info = [
            ("파견비용", df_dispatch),
            ("컨설팅", df_consulting),
            ("수입3팀", df_income_team3),
            ("수입2팀", df_income_team2),
            ("수출팀", df_export),
            ("매출원가", df_cost_of_sales),
            ("수입1팀", df_import_team1),
        ]
        for col, (name, team_df) in zip(cols, teams_info):
            col.metric(name, f"{len(team_df)}건", f"{fmt_money(team_df['공급가'].sum())}원")

        with st.expander("데이터 검증 (원본 vs 합계 비교)"):
            df_val = create_validation_sheet(df, df_dispatch, df_consulting, df_income_team3, df_income_team2, df_export, df_cost_of_sales, df_import_team1)
            diff = df_val[df_val['구분'] == '차이']['공급가'].values[0]
            if abs(diff) < 0.5:
                st.success("원본과 분류 합계가 정확히 일치합니다.")
            else:
                st.warning(f"차이 발생: {diff:,.0f}원")
            st.dataframe(
                df_val.style.format({"공급가": "{:,.0f}", "건수": "{:.0f}"}),
                hide_index=True,
                width="stretch"
            )

        summaries = display_dashboard(df_consulting, df_income_team3, df_income_team2, df_export, df_import_team1, config)

        st.markdown("---")
        st.subheader("결과 다운로드")
        col_dl1, col_dl2 = st.columns(2)

        base_name = os.path.splitext(uploaded_file.name)[0]

        with col_dl1:
            excel_data = to_excel(df, df_dispatch, df_consulting, df_income_team3, df_income_team2, df_export, df_cost_of_sales, df_import_team1, config)
            st.download_button(
                label="엑셀 파일 다운로드",
                data=excel_data,
                file_name=f"월정산_결과_{base_name}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch"
            )

        with col_dl2:
            report_cols = ['받는자', '과목명', '업무']
            extra_summaries = {
                '파견비용': create_full_summary(df_dispatch, group_cols=report_cols),
                '매출원가': create_full_summary(df_cost_of_sales, group_cols=report_cols),
            }
            html_report = to_html(summaries, config, extra_summaries=extra_summaries)
            st.download_button(
                label="HTML 보고서 다운로드",
                data=html_report,
                file_name=f"월정산_보고서_{base_name}.html",
                mime="text/html",
                width="stretch"
            )

else:
    st.info("원본 데이터 파일을 업로드해주세요. 설정은 좌측 사이드바에서 관리합니다.")
