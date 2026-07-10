import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 새 워크북 생성
wb = openpyxl.Workbook()

# 기본 시트 제거
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# 스타일 정의
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
info_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

def apply_header_style(ws, row_num, col_count):
    """헤더 스타일 적용"""
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

def apply_data_style(ws, start_row, end_row, start_col, end_col):
    """데이터 영역 스타일 적용"""
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='center')

def add_info_box(ws, row, message):
    """안내 메시지 박스 추가"""
    ws.merge_cells(f'A{row}:D{row}')
    cell = ws.cell(row=row, column=1)
    cell.value = message
    cell.fill = info_fill
    cell.font = Font(bold=True, size=10)
    cell.alignment = Alignment(horizontal='left', vertical='center')

# ============================================
# 1. 설정_수입2팀 시트
# ============================================
ws_import2 = wb.create_sheet("설정_수입2팀")

# 안내 메시지
add_info_box(ws_import2, 1, "📌 수입2팀 필터링 조건 | 행을 추가하거나 삭제하여 조건을 관리하세요")

# 헤더
ws_import2['A3'] = '받는자'
ws_import2['B3'] = '업무'
ws_import2['C3'] = '실화주'
ws_import2['D3'] = '비고'
apply_header_style(ws_import2, 3, 4)

# 데이터 (현재 코드의 shipper_conditions)
data_import2 = [
    ['(주)영화케이스틸', '수입', '', ''],
    ['에스엠월드와이드코리아(주)', '수입', '에스케이텔링크(주)', ''],
    ['큐리옥스바이오시스템즈(주)', '수출', '큐리옥스바이오시스템즈(주)', ''],
    ['(주)코리아인터링크-', '수출', '', ''],
    ['록키매니지먼트(주)-', '수출', '', ''],
    ['(주)피엔케이트레이딩', '수입', '', ''],
    ['(주)경일하이텍', '수입', '', ''],
    ['주식회사 엑트로지스틱스', '수입', '', ''],
    ['(주)에스더블유더블유로지스', '수출', '', ''],
    ['(주)노바미디어', '수입', '', ''],
    ['세미크론댄포스(주)', '수입', '', ''],
]

for idx, row_data in enumerate(data_import2, start=4):
    for col_idx, value in enumerate(row_data, start=1):
        ws_import2.cell(row=idx, column=col_idx, value=value)

apply_data_style(ws_import2, 4, 4 + len(data_import2) - 1, 1, 4)

# 열 너비 조정
ws_import2.column_dimensions['A'].width = 35
ws_import2.column_dimensions['B'].width = 12
ws_import2.column_dimensions['C'].width = 35
ws_import2.column_dimensions['D'].width = 20

# ============================================
# 2. 설정_수출팀 시트
# ============================================
ws_export = wb.create_sheet("설정_수출팀")

# 안내 메시지
add_info_box(ws_export, 1, "📌 수출팀 필터링 조건 | 등록자 명단 및 특정 업체 조건을 관리하세요")

# 섹션 1: 대상 업무 목록
ws_export['A3'] = '대상 업무 목록'
ws_export['B3'] = '비고'
ws_export.merge_cells('A3:A3')
apply_header_style(ws_export, 3, 2)

export_tasks = ['수출', '갈음', '환급']
for idx, task in enumerate(export_tasks, start=4):
    ws_export.cell(row=idx, column=1, value=task)
    ws_export.cell(row=idx, column=2, value='')

apply_data_style(ws_export, 4, 4 + len(export_tasks) - 1, 1, 2)

# 빈 행
ws_export.cell(row=10, column=1, value='')

# 섹션 2: 특정 업체 조건
ws_export['A11'] = '특정 업체 조건'
ws_export.merge_cells('A11:C11')
cell = ws_export['A11']
cell.fill = info_fill
cell.font = Font(bold=True, size=10)
cell.alignment = Alignment(horizontal='left', vertical='center')

ws_export['A12'] = '받는자'
ws_export['B12'] = '업무1'
ws_export['C12'] = '업무2'
ws_export['D12'] = '비고'
apply_header_style(ws_export, 12, 4)

ws_export.cell(row=13, column=1, value='삼성물산(주).')
ws_export.cell(row=13, column=2, value='수출')
ws_export.cell(row=13, column=3, value='환급')
ws_export.cell(row=13, column=4, value='')

apply_data_style(ws_export, 13, 13, 1, 4)

# 열 너비 조정
ws_export.column_dimensions['A'].width = 25
ws_export.column_dimensions['B'].width = 12
ws_export.column_dimensions['C'].width = 12
ws_export.column_dimensions['D'].width = 20

# ============================================
# 3. 설정_매출원가 시트
# ============================================
ws_cost = wb.create_sheet("설정_매출원가")

# 안내 메시지
add_info_box(ws_cost, 1, "📌 매출원가 과목명 목록 | 행을 추가하거나 삭제하여 과목명을 관리하세요")

# 헤더
ws_cost['A3'] = '과목명'
ws_cost['B3'] = '비고'
apply_header_style(ws_cost, 3, 2)

# 데이터
cost_categories = [
    '운송료',
    '경과보관료',
    '적출료',
    '창고료',
    '검역검사수수료',
    'H/C',
    '컨테이너적출료',
    '보세운송료',
    '작업료',
    '검역수수료',
    '보수작업료',
    '컨테이너 검사료',
    '검사수수료',
    '폐기물처리비용'
]

for idx, category in enumerate(cost_categories, start=4):
    ws_cost.cell(row=idx, column=1, value=category)
    ws_cost.cell(row=idx, column=2, value='')

apply_data_style(ws_cost, 4, 4 + len(cost_categories) - 1, 1, 2)

# 열 너비 조정
ws_cost.column_dimensions['A'].width = 25
ws_cost.column_dimensions['B'].width = 30

# ============================================
# 4. 설정_수입3팀 시트
# ============================================
ws_import3 = wb.create_sheet("설정_수입3팀")

# 안내 메시지
add_info_box(ws_import3, 1, "📌 수입3팀 필터링 조건 | 받는자 조건을 관리하세요 (업무구분 상관없음)")

# 헤더
ws_import3['A3'] = '받는자'
ws_import3['B3'] = '비고'
apply_header_style(ws_import3, 3, 2)

# 데이터
ws_import3.cell(row=4, column=1, value='삼성바이오로직스주식회사')
ws_import3.cell(row=4, column=2, value='최우선 조건')

apply_data_style(ws_import3, 4, 4, 1, 2)

# 열 너비 조정
ws_import3.column_dimensions['A'].width = 35
ws_import3.column_dimensions['B'].width = 30

# ============================================
# 5. 설정_컨설팅 시트
# ============================================
ws_consulting = wb.create_sheet("설정_컨설팅")

# 안내 메시지
add_info_box(ws_consulting, 1, "📌 컨설팅 필터링 조건 | 업무 조건을 관리하세요")

# 헤더
ws_consulting['A3'] = '업무'
ws_consulting['B3'] = '비고'
apply_header_style(ws_consulting, 3, 2)

# 데이터
ws_consulting.cell(row=4, column=1, value='기타')
ws_consulting.cell(row=4, column=2, value='')

apply_data_style(ws_consulting, 4, 4, 1, 2)

# 열 너비 조정
ws_consulting.column_dimensions['A'].width = 20
ws_consulting.column_dimensions['B'].width = 30

# ============================================
# 6. 사용 가이드 시트
# ============================================
ws_guide = wb.create_sheet("📖사용가이드", 0)  # 첫 번째 시트로 추가

ws_guide.merge_cells('A1:D1')
title_cell = ws_guide['A1']
title_cell.value = "월정산 시스템 - 설정 가이드"
title_cell.font = Font(bold=True, size=16, color="1F4E78")
title_cell.alignment = Alignment(horizontal='center', vertical='center')
ws_guide.row_dimensions[1].height = 30

# 개요
ws_guide['A3'] = "📋 개요"
ws_guide['A3'].font = Font(bold=True, size=12, color="1F4E78")
ws_guide['A4'] = "이 파일의 설정 시트들을 수정하여 각 팀의 필터링 조건을 관리할 수 있습니다."
ws_guide['A4'].alignment = Alignment(wrap_text=True)

# 처리 순서
ws_guide['A6'] = "⚙️ 데이터 처리 순서 (우선순위)"
ws_guide['A6'].font = Font(bold=True, size=12, color="1F4E78")

priority_info = [
    ["1순위", "컨설팅", "설정_컨설팅 시트의 조건 적용"],
    ["2순위", "수입3팀", "설정_수입3팀 시트의 조건 적용 (업무구분 무관)"],
    ["3순위", "매출원가", "설정_매출원가 시트의 과목명 조건 적용"],
    ["4순위", "수출팀", "설정_수출팀 시트의 조건 적용"],
    ["5순위", "수입2팀", "설정_수입2팀 시트의 조건 적용"],
    ["6순위", "수입1팀", "위 조건에 해당하지 않는 나머지 모든 데이터"],
]

ws_guide['A7'] = "우선순위"
ws_guide['B7'] = "팀명"
ws_guide['C7'] = "설명"
apply_header_style(ws_guide, 7, 3)

for idx, row_data in enumerate(priority_info, start=8):
    for col_idx, value in enumerate(row_data, start=1):
        ws_guide.cell(row=idx, column=col_idx, value=value)

apply_data_style(ws_guide, 8, 8 + len(priority_info) - 1, 1, 3)

# 사용 방법
ws_guide['A15'] = "📝 설정 시트 수정 방법"
ws_guide['A15'].font = Font(bold=True, size=12, color="1F4E78")

instructions = [
    "1. 조건 추가: 해당 설정 시트에서 새 행을 추가하고 값을 입력하세요",
    "2. 조건 삭제: 해당 행을 전체 삭제하세요",
    "3. 조건 수정: 셀의 값을 직접 수정하세요",
    "4. 빈 칸: '실화주'나 '비고' 등 선택 항목은 비워둘 수 있습니다",
]

for idx, instruction in enumerate(instructions, start=16):
    cell = ws_guide.cell(row=idx, column=1, value=instruction)
    cell.alignment = Alignment(wrap_text=True)

# 주의사항
ws_guide['A21'] = "⚠️ 주의사항"
ws_guide['A21'].font = Font(bold=True, size=12, color="C00000")

warnings = [
    "• 설정 시트의 이름을 변경하지 마세요",
    "• 헤더 행(3행)을 삭제하거나 수정하지 마세요",
    "• 원본 시트의 이름은 반드시 '원본'이어야 합니다",
    "• 조건이 중복되면 우선순위가 높은 팀으로 배정됩니다",
]

for idx, warning in enumerate(warnings, start=22):
    cell = ws_guide.cell(row=idx, column=1, value=warning)
    cell.font = Font(color="C00000")

# 열 너비 조정
ws_guide.column_dimensions['A'].width = 15
ws_guide.column_dimensions['B'].width = 15
ws_guide.column_dimensions['C'].width = 50

# 파일 저장
output_path = r"c:\Users\PC\OneDrive\Desktop\python\03-2.본사월정산-스트림릿\설정_템플릿_v2.xlsx"
wb.save(output_path)
print(f"✅ 설정 템플릿이 생성되었습니다: {output_path}")
